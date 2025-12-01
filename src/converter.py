"""
Konwerter sprawozdań finansowych do formatu XLSX.

Generuje wieloarkuszowy plik Excel:
1. Bilans - format czytelny
2. RZiS - format czytelny
3. Nota podatkowa - jeśli dostępna
4. Zestawienie zmian w kapitale - jeśli dostępne (jednostki Inna)
5. Rachunek przepływów - jeśli dostępny (jednostki Inna)
6. Analiza wskaźnikowa - wskaźniki niewypłacalności (tylko wersja lokalna)
7. Dane surowe - wszystkie pozycje z kodami
8. Dane analityczne - format długi do analizy
"""

from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from models import Sprawozdanie, PozycjaFinansowa


class XLSXConverter:
    """Konwerter sprawozdania do formatu XLSX."""

    # Style
    HEADER_FONT = Font(bold=True)
    HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    TITLE_FONT = Font(bold=True, size=12)
    MONEY_FORMAT = '#,##0.00'
    DATE_FORMAT = 'YYYY-MM-DD'

    # Style dla analizy wskaźnikowej
    OCENA_FILL = {
        "optymalna": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # jasny zielony
        "akceptowalna": PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid"),  # bardzo jasny zielony
        "ostrzegawcza": PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),  # jasny żółty
        "krytyczna": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),  # jasny czerwony
        "brak_danych": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),  # szary
    }

    def __init__(self):
        self.wb = None
        self.spr = None

    def convert(self, sprawozdanie: Sprawozdanie, output_dir: Path) -> tuple:
        """Konwertuje sprawozdanie do XLSX i zapisuje do pliku wraz z załącznikami.

        Args:
            sprawozdanie: Sparsowane sprawozdanie finansowe
            output_dir: Katalog wyjściowy

        Returns:
            Tuple (ścieżka_xlsx, lista_ścieżek_załączników)
        """
        self.spr = sprawozdanie
        self.wb = Workbook()

        # Arkusz 1: Bilans
        ws_bilans = self.wb.active
        ws_bilans.title = "Bilans"
        self._create_bilans_sheet(ws_bilans)

        # Arkusz 2: RZiS
        wariant = "w.por." if sprawozdanie.metadane.wariant_rzis == "porownawczy" else "w.kalk."
        ws_rzis = self.wb.create_sheet(f"RZiS ({wariant})")
        self._create_rzis_sheet(ws_rzis)

        # Arkusz 3: Nota podatkowa (jeśli dostępna)
        if sprawozdanie.nota_podatkowa:
            ws_nota = self.wb.create_sheet("Nota podatkowa")
            self._create_nota_sheet(ws_nota)

        # Arkusz 4: Zestawienie zmian w kapitale własnym (jeśli dostępne)
        if sprawozdanie.zestawienie_zmian_kapital:
            ws_kapital = self.wb.create_sheet("Zest. zmian w kapitale")
            self._create_kapital_sheet(ws_kapital)

        # Arkusz 5: Rachunek przepływów pieniężnych (jeśli dostępny)
        if sprawozdanie.rachunek_przeplywow:
            wariant = "bezp." if sprawozdanie.wariant_przeplywow == "bezposredni" else "pośr."
            ws_przeplywy = self.wb.create_sheet(f"Rach. przepływów ({wariant})")
            self._create_przeplywy_sheet(ws_przeplywy)

        # Arkusz 6: Analiza wskaźnikowa (tylko wersja lokalna)
        ws_wskazniki = self.wb.create_sheet("Analiza wskaźnikowa")
        self._create_indicators_sheet(ws_wskazniki)

        # Arkusz 7: Dane surowe
        ws_surowe = self.wb.create_sheet("Dane surowe")
        self._create_raw_data_sheet(ws_surowe)

        # Arkusz 8: Dane analityczne
        ws_analityczne = self.wb.create_sheet("Dane analityczne")
        self._create_analytical_sheet(ws_analityczne)

        # Generuj nazwę pliku i zapisz
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        filename = sprawozdanie.nazwa_pliku()
        output_path = output_dir / filename

        self.wb.save(output_path)

        # Zapisz załączniki binarne (jeśli istnieją)
        zalaczniki_paths = self._save_attachments(output_dir)

        return output_path, zalaczniki_paths

    def _create_bilans_sheet(self, ws):
        """Tworzy arkusz Bilans w formacie czytelnym."""
        meta = self.spr.metadane
        firma = self.spr.dane_firmy
        weryfikacja = self.spr.weryfikacja

        # Nagłówek
        ws['A1'] = "BILANS"
        ws['A1'].font = self.TITLE_FONT

        # Dane firmy
        ws['A2'] = "Firma:"
        ws['B2'] = firma.nazwa

        ws['A3'] = "NIP:"
        ws['B3'] = firma.nip

        if firma.krs:
            ws['A4'] = "KRS:"
            ws['B4'] = firma.krs

        # Okresy
        row = 6
        ws[f'A{row}'] = "Okres sprawozdawczy:"
        ws[f'B{row}'] = f"{meta.okres_od} - {meta.okres_do}"

        row += 1
        ws[f'A{row}'] = "Typ jednostki:"
        ws[f'B{row}'] = meta.typ_jednostki

        row += 1
        ws[f'A{row}'] = "Wersja schematu:"
        ws[f'B{row}'] = meta.wersja_schematu

        row += 1
        ws[f'A{row}'] = "Jednostka walutowa:"
        ws[f'B{row}'] = meta.jednostka_walutowa
        if meta.jednostka_walutowa == "tys. PLN":
            ws[f'B{row}'].font = Font(bold=True, color="0066CC")

        # Weryfikacja sum
        row += 2
        ws[f'A{row}'] = "Weryfikacja sum:"
        if weryfikacja:
            ws[f'B{row}'] = "Aktywa = Pasywa" if weryfikacja.aktywa_rowne_pasywom_biezacy else "BŁĄD: Aktywa ≠ Pasywa"
            if not weryfikacja.aktywa_rowne_pasywom_biezacy:
                ws[f'B{row}'].font = Font(color="FF0000", bold=True)

        # Nagłówki kolumn
        row += 2
        ws[f'A{row}'] = "Pozycja"
        jednostka_skrot = "tys." if meta.jednostka_walutowa == "tys. PLN" else ""
        ws[f'B{row}'] = f"Rok {meta.okres_do.year} [{meta.jednostka_walutowa}]"
        ws[f'C{row}'] = f"Rok {meta.okres_do.year - 1} [{meta.jednostka_walutowa}]"

        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = self.HEADER_FONT
            ws[f'{col}{row}'].fill = self.HEADER_FILL

        # AKTYWA
        row += 2
        ws[f'A{row}'] = "AKTYWA"
        ws[f'A{row}'].font = self.TITLE_FONT
        row += 1

        for poz in self.spr.bilans_aktywa:
            row = self._write_position_row(ws, row, poz)

        # PASYWA
        row += 2
        ws[f'A{row}'] = "PASYWA"
        ws[f'A{row}'].font = self.TITLE_FONT
        row += 1

        for poz in self.spr.bilans_pasywa:
            row = self._write_position_row(ws, row, poz)

        # Formatowanie kolumn
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18

    def _create_rzis_sheet(self, ws):
        """Tworzy arkusz RZiS w formacie czytelnym."""
        meta = self.spr.metadane
        firma = self.spr.dane_firmy

        # Nagłówek
        wariant_nazwa = "WARIANT PORÓWNAWCZY" if meta.wariant_rzis == "porownawczy" else "WARIANT KALKULACYJNY"
        ws['A1'] = f"RACHUNEK ZYSKÓW I STRAT ({wariant_nazwa})"
        ws['A1'].font = self.TITLE_FONT

        # Dane firmy
        ws['A2'] = "Firma:"
        ws['B2'] = firma.nazwa

        ws['A3'] = "Okres:"
        ws['B3'] = f"{meta.okres_od} - {meta.okres_do}"

        ws['A4'] = "Jednostka walutowa:"
        ws['B4'] = meta.jednostka_walutowa
        if meta.jednostka_walutowa == "tys. PLN":
            ws['B4'].font = Font(bold=True, color="0066CC")

        # Nagłówki kolumn
        row = 6
        ws[f'A{row}'] = "Pozycja"
        ws[f'B{row}'] = f"Rok {meta.okres_do.year} [{meta.jednostka_walutowa}]"
        ws[f'C{row}'] = f"Rok {meta.okres_do.year - 1} [{meta.jednostka_walutowa}]"

        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = self.HEADER_FONT
            ws[f'{col}{row}'].fill = self.HEADER_FILL

        # Pozycje RZiS
        row += 1
        for poz in self.spr.rzis:
            row = self._write_position_row(ws, row, poz)

        # Formatowanie kolumn
        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18

    def _create_nota_sheet(self, ws):
        """Tworzy arkusz Nota podatkowa."""
        meta = self.spr.metadane
        firma = self.spr.dane_firmy

        # Nagłówek
        ws['A1'] = "NOTA PODATKOWA"
        ws['A1'].font = self.TITLE_FONT

        ws['A2'] = "Firma:"
        ws['B2'] = firma.nazwa

        # Nagłówki kolumn
        row = 4
        ws[f'A{row}'] = "Pozycja"
        ws[f'B{row}'] = f"Rok {meta.okres_do.year}"
        ws[f'C{row}'] = f"Rok {meta.okres_do.year - 1}"

        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = self.HEADER_FONT
            ws[f'{col}{row}'].fill = self.HEADER_FILL

        # Pozycje
        row += 1
        for poz in self.spr.nota_podatkowa:
            row = self._write_position_row(ws, row, poz)

        # Formatowanie kolumn
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18

    def _create_kapital_sheet(self, ws):
        """Tworzy arkusz Zestawienie zmian w kapitale własnym."""
        meta = self.spr.metadane
        firma = self.spr.dane_firmy

        # Nagłówek
        ws['A1'] = "ZESTAWIENIE ZMIAN W KAPITALE (FUNDUSZU) WŁASNYM"
        ws['A1'].font = self.TITLE_FONT

        ws['A2'] = "Firma:"
        ws['B2'] = firma.nazwa

        ws['A3'] = "Okres:"
        ws['B3'] = f"{meta.okres_od} - {meta.okres_do}"

        # Nagłówki kolumn
        row = 5
        ws[f'A{row}'] = "Pozycja"
        ws[f'B{row}'] = f"Rok {meta.okres_do.year}"
        ws[f'C{row}'] = f"Rok {meta.okres_do.year - 1}"

        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = self.HEADER_FONT
            ws[f'{col}{row}'].fill = self.HEADER_FILL

        # Pozycje
        row += 1
        for poz in self.spr.zestawienie_zmian_kapital:
            row = self._write_position_row(ws, row, poz)

        # Formatowanie kolumn
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18

    def _create_przeplywy_sheet(self, ws):
        """Tworzy arkusz Rachunek przepływów pieniężnych."""
        meta = self.spr.metadane
        firma = self.spr.dane_firmy

        # Nagłówek
        wariant_nazwa = "METODA BEZPOŚREDNIA" if self.spr.wariant_przeplywow == "bezposredni" else "METODA POŚREDNIA"
        ws['A1'] = f"RACHUNEK PRZEPŁYWÓW PIENIĘŻNYCH ({wariant_nazwa})"
        ws['A1'].font = self.TITLE_FONT

        ws['A2'] = "Firma:"
        ws['B2'] = firma.nazwa

        ws['A3'] = "Okres:"
        ws['B3'] = f"{meta.okres_od} - {meta.okres_do}"

        # Nagłówki kolumn
        row = 5
        ws[f'A{row}'] = "Pozycja"
        ws[f'B{row}'] = f"Rok {meta.okres_do.year}"
        ws[f'C{row}'] = f"Rok {meta.okres_do.year - 1}"

        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = self.HEADER_FONT
            ws[f'{col}{row}'].fill = self.HEADER_FILL

        # Pozycje
        row += 1
        for poz in self.spr.rachunek_przeplywow:
            row = self._write_position_row(ws, row, poz)

        # Formatowanie kolumn
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18

    def _create_indicators_sheet(self, ws):
        """Tworzy arkusz z analizą wskaźnikową niewypłacalności.

        UWAGA: Ten arkusz jest dostępny TYLKO w wersji lokalnej aplikacji.
        """
        from indicators import (
            KalkulatorWskaznikow,
            extract_financial_data_from_sprawozdanie,
            OcenaWskaznika
        )

        meta = self.spr.metadane
        firma = self.spr.dane_firmy

        # Nagłówek
        ws['A1'] = "ANALIZA WSKAŹNIKOWA - OCENA NIEWYPŁACALNOŚCI"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A2'] = "Na podstawie: Meritum, B. Prusak, K. Prędkiewicz, Doradca Restrukturyzacyjny"
        ws['A2'].font = Font(italic=True, size=9)

        ws['A4'] = "Firma:"
        ws['B4'] = firma.nazwa

        ws['A5'] = "Okres:"
        ws['B5'] = f"{meta.okres_od} - {meta.okres_do}"

        ws['A6'] = "Typ jednostki:"
        ws['B6'] = meta.typ_jednostki

        ws['A7'] = "Jednostka walutowa:"
        ws['B7'] = meta.jednostka_walutowa
        if meta.jednostka_walutowa == "tys. PLN":
            ws['B7'].font = Font(bold=True, color="0066CC")

        # Ostrzeżenie
        ws['A9'] = "UWAGA: Wskaźniki pełnią funkcję pomocniczą. Pełna ocena wymaga analizy dynamicznej i kontekstu branżowego."
        ws['A9'].font = Font(italic=True, color="666666")

        # Wyciągnij dane finansowe i oblicz wskaźniki
        dane_finansowe = extract_financial_data_from_sprawozdanie(self.spr)
        kalkulator = KalkulatorWskaznikow(dane_finansowe)
        wyniki = kalkulator.oblicz_wszystkie()

        # Nagłówki tabeli
        row = 11
        headers = ["Wskaźnik", "Skrót", "Wzór", "Wartość", "Ocena", "Interpretacja", "Optimum", "Wart. krytyczna", "Źródło"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL

        # Dane wskaźników
        row += 1
        current_section = None

        for wynik in wyniki:
            # Dodaj separatory między sekcjami (na podstawie nazwy wskaźnika)
            section = self._get_indicator_section(wynik.nazwa)
            if section != current_section:
                if current_section is not None:
                    row += 1  # Pusta linia między sekcjami
                current_section = section

            ws.cell(row=row, column=1, value=wynik.nazwa)
            ws.cell(row=row, column=2, value=wynik.skrot)
            ws.cell(row=row, column=3, value=wynik.wzor)  # Nowa kolumna ze wzorem
            ws.cell(row=row, column=4, value=wynik.wartosc_str)
            ws.cell(row=row, column=5, value=wynik.ocena.value)
            ws.cell(row=row, column=6, value=wynik.interpretacja)
            ws.cell(row=row, column=7, value=wynik.optimum)
            ws.cell(row=row, column=8, value=wynik.wartosc_krytyczna)
            ws.cell(row=row, column=9, value=wynik.zrodlo)

            # Kolorowanie oceny
            ocena_cell = ws.cell(row=row, column=5)
            if wynik.ocena.value in self.OCENA_FILL:
                ocena_cell.fill = self.OCENA_FILL[wynik.ocena.value]

            # Kolorowanie wartości przy krytycznej ocenie
            if wynik.ocena == OcenaWskaznika.KRYTYCZNA:
                ws.cell(row=row, column=4).font = Font(bold=True, color="CC0000")

            row += 1

        # Podsumowanie ocen
        row += 2
        ws.cell(row=row, column=1, value="PODSUMOWANIE OCEN:")
        ws.cell(row=row, column=1).font = Font(bold=True)

        row += 1
        oceny_count = {
            "optymalna": 0,
            "akceptowalna": 0,
            "ostrzegawcza": 0,
            "krytyczna": 0,
            "brak_danych": 0,
        }
        for wynik in wyniki:
            oceny_count[wynik.ocena.value] += 1

        for ocena, count in oceny_count.items():
            ws.cell(row=row, column=1, value=f"{ocena.capitalize()}:")
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=1).fill = self.OCENA_FILL.get(ocena, PatternFill())
            row += 1

        # Sekcja z objaśnieniami wzorów i modeli
        row += 2
        ws.cell(row=row, column=1, value="OBJAŚNIENIA WZORÓW I SYMBOLI:")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

        row += 2
        ws.cell(row=row, column=1, value="WSKAŹNIKI PŁYNNOŚCI")
        ws.cell(row=row, column=1).font = Font(bold=True, underline="single")
        row += 1
        objasnienia_plynnosc = [
            ("CR (Current Ratio)", "Aktywa obrotowe / Zobowiązania krótkoterminowe", "Zdolność do pokrycia zobowiązań bieżących aktywami obrotowymi. Optymalnie: 1,3-2,0"),
            ("QR (Quick Ratio)", "(AO - Zapasy - RMK) / Zobowiązania krótkoterminowe", "Płynność po wyłączeniu najmniej płynnych aktywów. Optymalnie: 1,0-1,2"),
            ("CaR (Cash Ratio)", "Środki pieniężne / Zobowiązania krótkoterminowe", "Natychmiastowa zdolność płatnicza. Optymalnie: 0,1-0,2"),
            ("KP (Kapitał pracujący)", "Aktywa obrotowe - Zobowiązania krótkoterminowe", "Bufor bezpieczeństwa płynności. Powinien być dodatni"),
        ]
        for skrot, wzor, opis in objasnienia_plynnosc:
            ws.cell(row=row, column=1, value=skrot).font = Font(bold=True)
            ws.cell(row=row, column=2, value=wzor)
            ws.cell(row=row, column=4, value=opis)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="WSKAŹNIKI ZADŁUŻENIA")
        ws.cell(row=row, column=1).font = Font(bold=True, underline="single")
        row += 1
        objasnienia_zadluzenie = [
            ("WOZ", "Zobowiązania ogółem / Aktywa ogółem × 100%", "Udział kapitału obcego w finansowaniu. Optymalnie: 57-67%"),
            ("WZK", "Zobowiązania ogółem / Kapitał własny", "Relacja długu do kapitału własnego. Dla MSP dopuszczalne do 3:1"),
            ("UKW", "Kapitał własny / Aktywa ogółem × 100%", "Udział kapitału własnego. Powinien wynosić min. 33%"),
        ]
        for skrot, wzor, opis in objasnienia_zadluzenie:
            ws.cell(row=row, column=1, value=skrot).font = Font(bold=True)
            ws.cell(row=row, column=2, value=wzor)
            ws.cell(row=row, column=4, value=opis)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="WSKAŹNIKI RENTOWNOŚCI")
        ws.cell(row=row, column=1).font = Font(bold=True, underline="single")
        row += 1
        objasnienia_rentownosc = [
            ("ROA", "Zysk netto / Aktywa ogółem × 100%", "Efektywność wykorzystania majątku"),
            ("ROS", "Zysk netto / Przychody ze sprzedaży × 100%", "Marża zysku netto na sprzedaży"),
            ("ROE", "Zysk netto / Kapitał własny × 100%", "Stopa zwrotu dla właścicieli"),
        ]
        for skrot, wzor, opis in objasnienia_rentownosc:
            ws.cell(row=row, column=1, value=skrot).font = Font(bold=True)
            ws.cell(row=row, column=2, value=wzor)
            ws.cell(row=row, column=4, value=opis)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="WSKAŹNIKI AKTYWNOŚCI")
        ws.cell(row=row, column=1).font = Font(bold=True, underline="single")
        row += 1
        objasnienia_aktywnosc = [
            ("CZ", "(Średnie zapasy / Przychody) × 365", "Okres utrzymywania zapasów w dniach"),
            ("CN", "(Średnie należności / Przychody) × 365", "Okres ściągania należności w dniach"),
            ("CZob", "(Średnie zobowiązania krótkoterm. / Przychody) × 365", "Okres regulowania zobowiązań w dniach"),
            ("CKG", "CZ + CN - CZob", "Cykl konwersji gotówki. Ujemny = finansowanie z kredytu kupieckiego"),
        ]
        for skrot, wzor, opis in objasnienia_aktywnosc:
            ws.cell(row=row, column=1, value=skrot).font = Font(bold=True)
            ws.cell(row=row, column=2, value=wzor)
            ws.cell(row=row, column=4, value=opis)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="MODELE DYSKRYMINACYJNE - POLSKIE")
        ws.cell(row=row, column=1).font = Font(bold=True, underline="single")
        row += 1
        objasnienia_modele_pl = [
            ("Model poznański (FD_P)", "3,562×X₁ + 1,588×X₂ + 4,288×X₃ + 6,719×X₄ - 2,368",
             "X₁=ZN/A, X₂=(AO-Zap)/ZK, X₃=KS/A, X₄=WS/PS. Próg: Z>0 brak zagrożenia. Trafność: 96%"),
            ("Model Prusaka 1r (FD_PR1)", "6,9973×X₁ + 0,1191×X₂ + 0,1932×X₃ - 1,1760",
             "X₁=WS/śrA, X₂=KO/śrZK, X₃=AO/ZK. Dedykowany MSP. Próg: Z>0"),
            ("Model Prusaka 2l (FD_PR2)", "3,7657×X₁ + 0,1049×X₂ - 1,6765×X₃ + 3,5230×X₄ - 0,3758",
             "Wersja z 2-letnim wyprzedzeniem. Próg: Z>0"),
            ("Model A. Hołdy (FD_H)", "0,605 + 0,681×X₁ - 0,0196×X₂ + 0,157×X₃ + 0,00969×X₄ + 0,000672×X₅",
             "Progi: Z≥0,1 brak zagrożenia, Z≤-0,3 zagrożenie, strefa szara pomiędzy"),
            ("Model Gajdki-Stosa (FD_GS)", "0,773 - 0,086×X₁ - 0,0008×X₂ + 0,922×X₃ + 0,654×X₄ - 0,595×X₅",
             "Próg: Z>0,45 dobra kondycja"),
            ("Model Hadasik (FD_HD)", "0,336×X₁ - 0,712×X₂ - 2,472×X₃ + 1,464×X₄ + 0,002×X₅ - 0,014×X₆ + 0,002×X₇ + 2,593",
             "Model 7-zmiennowy. Próg: Z≥0 brak zagrożenia"),
            ("Model Mączyńskiej (FD_M)", "1,50×X₁ + 0,08×X₂ + 10,00×X₃ + 5,00×X₄ + 0,30×X₅ + 0,10×X₆",
             "Progi: Z≥1 dobra kondycja, 0<Z<1 słaba, Z≤0 zagrożenie"),
            ("Model Wierzby (FD_W)", "3,26×X₁ + 2,16×X₂ + 0,30×X₃ + 0,69×X₄",
             "Próg: Z>0 dobra kondycja"),
        ]
        for skrot, wzor, opis in objasnienia_modele_pl:
            ws.cell(row=row, column=1, value=skrot).font = Font(bold=True)
            ws.cell(row=row, column=2, value=wzor)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            row += 1
            ws.cell(row=row, column=2, value=opis).font = Font(italic=True, size=9)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="MODELE DYSKRYMINACYJNE - ZAGRANICZNE")
        ws.cell(row=row, column=1).font = Font(bold=True, underline="single")
        row += 1
        objasnienia_modele_zagr = [
            ("Model Altmana (FD_A)", "1,2×X₁ + 1,4×X₂ + 3,3×X₃ + 0,6×X₄ + 1,0×X₅",
             "X₁=KP/A, X₂=ZZ/A, X₃=EBIT/A, X₄=KW/ZO, X₅=PS/A. Progi: Z≥3 bezpieczna, 1,8<Z<3 strefa szara, Z≤1,8 zagrożenie"),
            ("Wilcox-Gambler (WL)", "ŚP + 0,70×Nal + 0,50×Zap + 0,50×Inne + 0,50×AT - ZK - ZD",
             "Wartość likwidacyjna majątku. WL>0 wypłacalność w ujęciu likwidacyjnym"),
        ]
        for skrot, wzor, opis in objasnienia_modele_zagr:
            ws.cell(row=row, column=1, value=skrot).font = Font(bold=True)
            ws.cell(row=row, column=2, value=wzor)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            row += 1
            ws.cell(row=row, column=2, value=opis).font = Font(italic=True, size=9)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="LEGENDA SYMBOLI:")
        ws.cell(row=row, column=1).font = Font(bold=True, underline="single")
        row += 1
        symbole = [
            "A = Aktywa ogółem | AO = Aktywa obrotowe | AT = Aktywa trwałe | KS = Kapitał stały (KW + ZD)",
            "KW = Kapitał własny | KP = Kapitał pracujący | KO = Koszty operacyjne",
            "ZO = Zobowiązania ogółem | ZK = Zobowiązania krótkoterminowe | ZD = Zobowiązania długoterminowe",
            "PS = Przychody netto ze sprzedaży | WS = Wynik ze sprzedaży | ZN = Zysk netto | ZB = Zysk brutto",
            "Zap = Zapasy | Nal = Należności | ŚP = Środki pieniężne | RMK = Rozliczenia międzyokresowe",
            "śr = średni stan (początek + koniec okresu) / 2",
        ]
        for sym in symbole:
            ws.cell(row=row, column=1, value=sym).font = Font(size=9)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            row += 1

        # Formatowanie kolumn (zaktualizowane dla dodatkowej kolumny Wzór)
        ws.column_dimensions['A'].width = 35  # Wskaźnik
        ws.column_dimensions['B'].width = 10  # Skrót
        ws.column_dimensions['C'].width = 55  # Wzór
        ws.column_dimensions['D'].width = 15  # Wartość
        ws.column_dimensions['E'].width = 12  # Ocena
        ws.column_dimensions['F'].width = 65  # Interpretacja
        ws.column_dimensions['G'].width = 20  # Optimum
        ws.column_dimensions['H'].width = 22  # Wart. krytyczna
        ws.column_dimensions['I'].width = 30  # Źródło

    def _get_indicator_section(self, nazwa: str) -> str:
        """Określa sekcję wskaźnika na podstawie nazwy."""
        nazwa_lower = nazwa.lower()
        if "płynność" in nazwa_lower or "gotówk" in nazwa_lower or "wystarczalność" in nazwa_lower:
            return "plynnosc"
        elif "zadłużeni" in nazwa_lower or "pokryci" in nazwa_lower:
            return "zadluzenie"
        elif "rentowność" in nazwa_lower:
            return "rentownosc"
        elif "model" in nazwa_lower or "wilcox" in nazwa_lower:
            return "modele"
        else:
            return "strukturalne"

    def _create_raw_data_sheet(self, ws):
        """Tworzy arkusz z danymi surowymi (wszystkie pozycje z kodami)."""
        # Nagłówki
        headers = ["sekcja", "kod", "opis", "rok_biezacy", "rok_poprzedni"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL

        # Dane
        row = 2
        wszystkie_pozycje = self.spr.wszystkie_pozycje()

        for poz in wszystkie_pozycje:
            ws.cell(row=row, column=1, value=poz.sekcja)
            ws.cell(row=row, column=2, value=poz.kod)
            ws.cell(row=row, column=3, value=poz.opis)

            if poz.kwota_biezaca is not None:
                cell = ws.cell(row=row, column=4, value=float(poz.kwota_biezaca))
                cell.number_format = self.MONEY_FORMAT

            if poz.kwota_poprzednia is not None:
                cell = ws.cell(row=row, column=5, value=float(poz.kwota_poprzednia))
                cell.number_format = self.MONEY_FORMAT

            row += 1

        # Formatowanie kolumn
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        # Włącz autofiltr
        ws.auto_filter.ref = f"A1:E{row-1}"

    def _create_analytical_sheet(self, ws):
        """Tworzy arkusz z danymi analitycznymi (format długi)."""
        meta = self.spr.metadane
        firma = self.spr.dane_firmy

        # Nagłówki
        headers = [
            "firma", "nip", "krs", "typ_jednostki", "wersja",
            "okres", "sekcja", "kod", "kod_pelny", "opis", "kwota"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL

        row = 2
        wszystkie_pozycje = self.spr.wszystkie_pozycje()

        for poz in wszystkie_pozycje:
            kod_pelny = poz.kod_pelny(meta.typ_jednostki, meta.wersja_schematu)

            # Rok bieżący
            if poz.kwota_biezaca is not None:
                ws.cell(row=row, column=1, value=firma.nazwa)
                ws.cell(row=row, column=2, value=firma.nip)
                ws.cell(row=row, column=3, value=firma.krs or "")
                ws.cell(row=row, column=4, value=meta.typ_jednostki)
                ws.cell(row=row, column=5, value=meta.wersja_schematu)
                ws.cell(row=row, column=6, value=meta.okres_do)
                ws.cell(row=row, column=7, value=poz.sekcja)
                ws.cell(row=row, column=8, value=poz.kod)
                ws.cell(row=row, column=9, value=kod_pelny)
                ws.cell(row=row, column=10, value=poz.opis)
                cell = ws.cell(row=row, column=11, value=float(poz.kwota_biezaca))
                cell.number_format = self.MONEY_FORMAT
                row += 1

            # Rok poprzedni
            if poz.kwota_poprzednia is not None:
                okres_poprz = date(meta.okres_do.year - 1, 12, 31)
                ws.cell(row=row, column=1, value=firma.nazwa)
                ws.cell(row=row, column=2, value=firma.nip)
                ws.cell(row=row, column=3, value=firma.krs or "")
                ws.cell(row=row, column=4, value=meta.typ_jednostki)
                ws.cell(row=row, column=5, value=meta.wersja_schematu)
                ws.cell(row=row, column=6, value=okres_poprz)
                ws.cell(row=row, column=7, value=poz.sekcja)
                ws.cell(row=row, column=8, value=poz.kod)
                ws.cell(row=row, column=9, value=kod_pelny)
                ws.cell(row=row, column=10, value=poz.opis)
                cell = ws.cell(row=row, column=11, value=float(poz.kwota_poprzednia))
                cell.number_format = self.MONEY_FORMAT
                row += 1

        # Formatowanie kolumn
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 35
        ws.column_dimensions['J'].width = 50
        ws.column_dimensions['K'].width = 15

        # Włącz autofiltr
        ws.auto_filter.ref = f"A1:K{row-1}"

    def _save_attachments(self, output_dir: Path) -> list:
        """Zapisuje załączniki binarne do katalogu wyjściowego.

        Args:
            output_dir: Katalog wyjściowy

        Returns:
            Lista ścieżek do zapisanych plików
        """
        saved_paths = []

        if not self.spr.zalaczniki:
            return saved_paths

        # Utwórz podkatalog dla załączników (z nazwą firmy)
        nazwa_clean = self.spr.dane_firmy.nazwa
        niedozwolone = '<>:"/\\|?*'
        nazwa_clean = "".join(c for c in nazwa_clean if c not in niedozwolone)
        if len(nazwa_clean) > 50:
            nazwa_clean = nazwa_clean[:50]

        attachments_dir = output_dir / f"zalaczniki_{nazwa_clean}"
        attachments_dir.mkdir(parents=True, exist_ok=True)

        for i, zal in enumerate(self.spr.zalaczniki, 1):
            # Unikalna nazwa pliku (unikamy nadpisywania)
            nazwa_pliku = zal.nazwa_pliku

            # Sprawdź czy plik już istnieje - jeśli tak, dodaj numer
            output_path = attachments_dir / nazwa_pliku
            if output_path.exists():
                base_name = nazwa_pliku.rsplit(".", 1)[0] if "." in nazwa_pliku else nazwa_pliku
                ext = f".{nazwa_pliku.rsplit('.', 1)[1]}" if "." in nazwa_pliku else ""
                output_path = attachments_dir / f"{base_name}_{i}{ext}"

            # Zapisz plik binarny
            try:
                with open(output_path, "wb") as f:
                    f.write(zal.zawartosc)
                saved_paths.append(output_path)
            except IOError as e:
                print(f"Błąd zapisu załącznika {nazwa_pliku}: {e}")

        return saved_paths

    def _write_position_row(self, ws, row: int, poz: PozycjaFinansowa) -> int:
        """Zapisuje wiersz pozycji finansowej.

        Args:
            ws: Arkusz
            row: Numer wiersza
            poz: Pozycja finansowa

        Returns:
            Numer następnego wiersza
        """
        # Wcięcie na podstawie poziomu
        indent = "  " * poz.poziom
        ws[f'A{row}'] = f"{indent}{poz.opis}"

        # Pogrubienie dla pozycji głównych
        if poz.poziom == 0:
            ws[f'A{row}'].font = Font(bold=True)

        # Kwoty
        if poz.kwota_biezaca is not None:
            cell = ws[f'B{row}']
            cell.value = float(poz.kwota_biezaca)
            cell.number_format = self.MONEY_FORMAT
            cell.alignment = Alignment(horizontal='right')

        if poz.kwota_poprzednia is not None:
            cell = ws[f'C{row}']
            cell.value = float(poz.kwota_poprzednia)
            cell.number_format = self.MONEY_FORMAT
            cell.alignment = Alignment(horizontal='right')

        return row + 1


def convert_file(xml_path: str, output_dir: str) -> tuple:
    """Funkcja pomocnicza do konwersji pojedynczego pliku.

    Args:
        xml_path: Ścieżka do pliku XML
        output_dir: Katalog wyjściowy

    Returns:
        Tuple (ścieżka_xlsx, lista_ścieżek_załączników)
    """
    from parser import SFParser

    parser = SFParser()
    sprawozdanie = parser.parse(Path(xml_path))

    converter = XLSXConverter()
    xlsx_path, attachments = converter.convert(sprawozdanie, Path(output_dir))

    return str(xlsx_path), [str(p) for p in attachments]
