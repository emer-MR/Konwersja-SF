"""
Konwerter wieloletni sprawozdań finansowych.

Łączy kilka sprawozdań tego samego podmiotu (ten sam NIP/KRS) w jeden plik
XLSX, w którym każdy rok jest osobną kolumną - umożliwia analizę porównawczą
"rok obok roku".

Używany przez batch.py, gdy dla jednego podmiotu dostępne są 2 lub więcej
sprawozdania. Dla pojedynczego sprawozdania batch.py korzysta z klasycznego
XLSXConverter (converter.py).

Arkusze pliku wynikowego:
1. Podsumowanie       - dane podmiotu, lista źródłowych sprawozdań, ostrzeżenia
2. Bilans            - aktywa i pasywa, kolumny = lata
3. RZiS              - rachunek zysków i strat, kolumny = lata
4. Nota podatkowa    - jeśli dostępna choć w jednym sprawozdaniu
5. Zest. zmian w kapitale - jeśli dostępne
6. Rach. przepływów  - jeśli dostępny
7. Analiza wskaźnikowa - wskaźniki niewypłacalności w ujęciu wieloletnim
8. Dane surowe       - wszystkie pozycje z kodami, kolumny = lata
9. Dane analityczne  - format długi, wszystkie lata i typy okresów
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from models import Sprawozdanie


def _rok(spr: Sprawozdanie) -> int:
    """Rok sprawozdawczy = rok daty końca okresu."""
    return spr.metadane.okres_do.year


def _oczysc_nazwe(nazwa: str, limit: int = 80) -> str:
    """Usuwa znaki niedozwolone w nazwach plików Windows."""
    niedozwolone = '<>:"/\\|?*'
    czysta = "".join(c for c in (nazwa or "") if c not in niedozwolone).strip()
    return czysta[:limit] if czysta else "podmiot"


class MultiYearConverter:
    """Łączy wiele sprawozdań jednego podmiotu w jeden wieloletni plik XLSX."""

    HEADER_FONT = Font(bold=True)
    HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    TITLE_FONT = Font(bold=True, size=12)
    SUBTITLE_FONT = Font(bold=True, size=11)
    BIG_TITLE_FONT = Font(bold=True, size=14)
    SECTION_FONT = Font(bold=True, underline="single")
    ITALIC_GREY = Font(italic=True, color="888888")
    WARN_FONT = Font(bold=True, color="CC0000")
    MONEY_FORMAT = '#,##0.00'

    OCENA_FILL = {
        "optymalna": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "akceptowalna": PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid"),
        "ostrzegawcza": PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),
        "krytyczna": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
        "brak_danych": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
    }

    SECTION_LABELS = {
        "plynnosc": "WSKAŹNIKI PŁYNNOŚCI",
        "zadluzenie": "WSKAŹNIKI ZADŁUŻENIA",
        "rentownosc": "WSKAŹNIKI RENTOWNOŚCI",
        "aktywnosc": "WSKAŹNIKI AKTYWNOŚCI I OBROTOWOŚCI",
        "strukturalne": "WSKAŹNIKI STRUKTURALNE",
        "modele": "MODELE DYSKRYMINACYJNE (PROGNOZA BANKRUCTWA)",
    }

    def __init__(self):
        self.wb = None
        self.reports = []       # list[Sprawozdanie] - posortowane rosnąco po roku
        self.sciezki = []       # list[Path] - równolegle do self.reports
        self.ostrzezenia = []   # list[str]
        self.jednostka_label = "PLN"

    # ------------------------------------------------------------------ API

    def convert(self, pary, output_dir) -> tuple:
        """Konwertuje grupę sprawozdań jednego podmiotu do wieloletniego XLSX.

        Args:
            pary: list[tuple[Path, Sprawozdanie]] - pliki źródłowe i sprawozdania
            output_dir: katalog wyjściowy

        Returns:
            tuple (ścieżka_xlsx, lista_ścieżek_załączników)
        """
        pary_sorted = sorted(
            pary,
            key=lambda ps: (ps[1].metadane.okres_do, ps[1].metadane.wersja_schematu),
        )
        self.reports = [s for _, s in pary_sorted]
        self.sciezki = [p for p, _ in pary_sorted]
        self.jednostka_label = self.reports[-1].metadane.jednostka_walutowa
        self.ostrzezenia = self._zbierz_ostrzezenia()

        self.wb = Workbook()

        ws_pods = self.wb.active
        ws_pods.title = "Podsumowanie"
        self._create_summary_sheet(ws_pods)

        self._create_financial_sheet(
            self.wb.create_sheet("Bilans"),
            "BILANS - UJĘCIE WIELOLETNIE",
            [("AKTYWA", lambda s: s.bilans_aktywa),
             ("PASYWA", lambda s: s.bilans_pasywa)],
        )

        self._create_financial_sheet(
            self.wb.create_sheet("RZiS"),
            "RACHUNEK ZYSKÓW I STRAT - UJĘCIE WIELOLETNIE",
            [(None, lambda s: s.rzis)],
        )

        if any(s.nota_podatkowa for s in self.reports):
            self._create_financial_sheet(
                self.wb.create_sheet("Nota podatkowa"),
                "NOTA PODATKOWA - UJĘCIE WIELOLETNIE",
                [(None, lambda s: s.nota_podatkowa)],
            )

        if any(s.zestawienie_zmian_kapital for s in self.reports):
            self._create_financial_sheet(
                self.wb.create_sheet("Zest. zmian w kapitale"),
                "ZESTAWIENIE ZMIAN W KAPITALE WŁASNYM - UJĘCIE WIELOLETNIE",
                [(None, lambda s: s.zestawienie_zmian_kapital)],
            )

        if any(s.rachunek_przeplywow for s in self.reports):
            self._create_financial_sheet(
                self.wb.create_sheet("Rach. przepływów"),
                "RACHUNEK PRZEPŁYWÓW PIENIĘŻNYCH - UJĘCIE WIELOLETNIE",
                [(None, lambda s: s.rachunek_przeplywow)],
            )

        self._create_indicators_sheet(self.wb.create_sheet("Analiza wskaźnikowa"))
        self._create_raw_sheet(self.wb.create_sheet("Dane surowe"))
        self._create_analytical_sheet(self.wb.create_sheet("Dane analityczne"))

        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / self._nazwa_pliku()
        self.wb.save(output_path)

        zalaczniki = self._save_attachments(output_dir)
        return output_path, zalaczniki

    # ------------------------------------------------------- nazwa / ostrzeżenia

    def _nazwa_pliku(self) -> str:
        lata = [_rok(s) for s in self.reports]
        nazwa = _oczysc_nazwe(self.reports[-1].dane_firmy.nazwa)
        return f"{min(lata)}-{max(lata)}_analiza-wieloletnia_{nazwa}.xlsx"

    def _zbierz_ostrzezenia(self) -> list:
        o = []
        waluty = {s.metadane.jednostka_walutowa for s in self.reports}
        if len(waluty) > 1:
            o.append(
                "UWAGA: sprawozdania mają różne jednostki walutowe "
                f"({', '.join(sorted(waluty))}) - kolumny lat moga byc nieporownywalne."
            )
        warianty = {s.metadane.wariant_rzis for s in self.reports}
        if len(warianty) > 1:
            o.append(
                "UWAGA: różne warianty RZiS (porównawczy / kalkulacyjny) - "
                "pozycje rachunku zyskow i strat moga sie roznic miedzy latami."
            )
        typy = {s.metadane.typ_jednostki for s in self.reports}
        if len(typy) > 1:
            o.append(
                f"UWAGA: różne typy jednostki ({', '.join(sorted(typy))}) - "
                "zakres prezentowanych pozycji moze sie roznic miedzy latami."
            )
        lata = [_rok(s) for s in self.reports]
        duplikaty = sorted({y for y in lata if lata.count(y) > 1})
        if duplikaty:
            o.append(
                f"UWAGA: kilka sprawozdań dla tego samego roku ({', '.join(map(str, duplikaty))}) "
                "- w kolumnie roku uzyto wartosci z ostatniego (najnowszy schemat)."
            )
        return o

    # ----------------------------------------------------------- łączenie danych

    def _merge_section(self, getter):
        """Łączy jedną sekcję (np. bilans_aktywa) ze wszystkich sprawozdań.

        Args:
            getter: funkcja Sprawozdanie -> list[PozycjaFinansowa] | None

        Returns:
            tuple (lata, wiersze):
              lata    - posortowana lista lat, dla których są dane
              wiersze - list[dict] z kluczami: kod, opis, poziom, sekcja,
                        values (dict rok->Decimal), only_old (bool)
        """
        # Kolejność wierszy bierzemy z najnowszego sprawozdania (wzorzec).
        wzorzec = getter(self.reports[-1]) or []
        kody_wzorca = [p.kod for p in wzorzec]
        widziane = set(kody_wzorca)

        # Pozycje obecne tylko w starszych sprawozdaniach (dopisywane na końcu).
        kody_dodatkowe = []
        for spr in self.reports:
            for p in getter(spr) or []:
                if p.kod not in widziane:
                    widziane.add(p.kod)
                    kody_dodatkowe.append(p.kod)

        # Opis i poziom - z najnowszego sprawozdania, które zawiera dany kod.
        meta = {}
        for spr in reversed(self.reports):
            for p in getter(spr) or []:
                if p.kod not in meta:
                    meta[p.kod] = (p.opis, p.poziom, p.sekcja)

        wszystkie_kody = kody_wzorca + kody_dodatkowe
        values = {kod: {} for kod in wszystkie_kody}

        # Najpierw dane porównawcze (rok-1) - niższy priorytet.
        for spr in self.reports:
            rok = _rok(spr)
            for p in getter(spr) or []:
                if p.kwota_poprzednia is not None:
                    values[p.kod].setdefault(rok - 1, p.kwota_poprzednia)

        # Następnie dane bieżące - zawsze nadpisują.
        for spr in self.reports:
            rok = _rok(spr)
            for p in getter(spr) or []:
                if p.kwota_biezaca is not None:
                    values[p.kod][rok] = p.kwota_biezaca

        lata = sorted({rok for kod in values for rok in values[kod]})

        wiersze = []
        for kod in kody_wzorca:
            opis, poziom, sekcja = meta[kod]
            wiersze.append(dict(kod=kod, opis=opis, poziom=poziom,
                                sekcja=sekcja, values=values[kod], only_old=False))
        for kod in kody_dodatkowe:
            opis, poziom, sekcja = meta[kod]
            wiersze.append(dict(kod=kod, opis=opis, poziom=poziom,
                                sekcja=sekcja, values=values[kod], only_old=True))
        return lata, wiersze

    # --------------------------------------------------------------- arkusze

    def _create_summary_sheet(self, ws):
        firma = self.reports[-1].dane_firmy
        lata = [_rok(s) for s in self.reports]

        ws['A1'] = "ANALIZA WIELOLETNIA SPRAWOZDAŃ FINANSOWYCH"
        ws['A1'].font = self.BIG_TITLE_FONT

        row = 3
        dane_podmiotu = [
            ("Podmiot:", firma.nazwa),
            ("NIP:", firma.nip or "-"),
            ("KRS:", firma.krs or "-"),
            ("REGON:", firma.regon or "-"),
            ("Adres:", firma.adres_pelny() or "-"),
            ("Liczba sprawozdań:", str(len(self.reports))),
            ("Zakres lat:", f"{min(lata)} - {max(lata)}"),
        ]
        for etykieta, wartosc in dane_podmiotu:
            ws.cell(row=row, column=1, value=etykieta).font = self.HEADER_FONT
            ws.cell(row=row, column=2, value=wartosc)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="ŹRÓDŁOWE SPRAWOZDANIA").font = self.SUBTITLE_FONT
        row += 1
        naglowki = ["Rok", "Okres", "Typ jednostki", "Wersja schematu",
                    "Wariant RZiS", "Jednostka walut.", "Plik źródłowy"]
        for col, h in enumerate(naglowki, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = self.HEADER_FONT
            c.fill = self.HEADER_FILL
        row += 1
        for spr, sciezka in zip(self.reports, self.sciezki):
            m = spr.metadane
            ws.cell(row=row, column=1, value=_rok(spr))
            ws.cell(row=row, column=2, value=f"{m.okres_od} - {m.okres_do}")
            ws.cell(row=row, column=3, value=m.typ_jednostki)
            ws.cell(row=row, column=4, value=m.wersja_schematu)
            ws.cell(row=row, column=5, value=m.wariant_rzis)
            ws.cell(row=row, column=6, value=m.jednostka_walutowa)
            ws.cell(row=row, column=7, value=Path(sciezka).name)
            row += 1

        if self.ostrzezenia:
            row += 1
            ws.cell(row=row, column=1, value="OSTRZEŻENIA").font = self.SUBTITLE_FONT
            row += 1
            for ostrz in self.ostrzezenia:
                c = ws.cell(row=row, column=1, value=ostrz)
                c.font = self.WARN_FONT
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                row += 1

        row += 1
        nota = ws.cell(
            row=row, column=1,
            value="Kwoty w arkuszach prezentowane są w jednostce walutowej "
                  f"najnowszego sprawozdania: {self.jednostka_label}.",
        )
        nota.font = Font(italic=True, color="666666")

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 30
        for col in 'CDEFG':
            ws.column_dimensions[col].width = 20
        ws.column_dimensions['G'].width = 48

    def _create_financial_sheet(self, ws, tytul, bloki):
        """Tworzy arkusz finansowy z kolumnami lat.

        Args:
            ws: arkusz
            tytul: tytuł arkusza
            bloki: list[tuple[str|None, getter]] - podtytuł sekcji + funkcja getter
        """
        firma = self.reports[-1].dane_firmy

        ws['A1'] = tytul
        ws['A1'].font = self.TITLE_FONT
        ws['A2'] = "Podmiot:"
        ws['A2'].font = self.HEADER_FONT
        ws['B2'] = firma.nazwa
        ws['A3'] = "NIP:"
        ws['A3'].font = self.HEADER_FONT
        ws['B3'] = firma.nip or "-"

        merged = [(sub, *self._merge_section(getter)) for sub, getter in bloki]
        lata = sorted({rok for _, lata_b, _ in merged for rok in lata_b})

        if not lata:
            ws['A5'] = "(brak danych w tej sekcji)"
            return

        header_row = 5
        c = ws.cell(row=header_row, column=1, value="Pozycja")
        c.font = self.HEADER_FONT
        c.fill = self.HEADER_FILL
        for i, rok in enumerate(lata):
            c = ws.cell(row=header_row, column=2 + i,
                        value=f"{rok} [{self.jednostka_label}]")
            c.font = self.HEADER_FONT
            c.fill = self.HEADER_FILL
            c.alignment = Alignment(horizontal='right')

        row = header_row + 1
        for sub, _lata_b, wiersze in merged:
            if sub:
                ws.cell(row=row, column=1, value=sub).font = self.SUBTITLE_FONT
                row += 1
            zwykle = [w for w in wiersze if not w['only_old']]
            stare = [w for w in wiersze if w['only_old']]
            for w in zwykle:
                row = self._write_pozycja_row(ws, row, w, lata)
            if stare:
                c = ws.cell(row=row, column=1,
                            value="Pozycje występujące tylko w starszych sprawozdaniach:")
                c.font = self.ITALIC_GREY
                row += 1
                for w in stare:
                    row = self._write_pozycja_row(ws, row, w, lata)
            row += 1

        ws.column_dimensions['A'].width = 65
        for i in range(len(lata)):
            ws.column_dimensions[get_column_letter(2 + i)].width = 18
        ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    def _write_pozycja_row(self, ws, row, wiersz, lata):
        indent = "  " * wiersz['poziom']
        cell = ws.cell(row=row, column=1, value=f"{indent}{wiersz['opis']}")
        if wiersz['poziom'] == 0:
            cell.font = Font(bold=True)
        for i, rok in enumerate(lata):
            wartosc = wiersz['values'].get(rok)
            if wartosc is not None:
                c = ws.cell(row=row, column=2 + i, value=float(wartosc))
                c.number_format = self.MONEY_FORMAT
                c.alignment = Alignment(horizontal='right')
        return row + 1

    def _create_indicators_sheet(self, ws):
        """Arkusz analizy wskaźnikowej - wskaźniki w kolumnach lat."""
        from indicators import (
            KalkulatorWskaznikow,
            extract_financial_data_from_sprawozdanie,
        )

        per_report = []  # list[tuple[rok, list[WynikWskaznika]]]
        for spr in self.reports:
            try:
                dane = extract_financial_data_from_sprawozdanie(spr)
                wyniki = KalkulatorWskaznikow(dane).oblicz_wszystkie()
            except Exception:
                wyniki = []
            per_report.append((_rok(spr), wyniki))

        ws['A1'] = "ANALIZA WSKAŹNIKOWA - UJĘCIE WIELOLETNIE"
        ws['A1'].font = self.BIG_TITLE_FONT
        ws['A2'] = ("Ocena niewypłacalności: płynność, zadłużenie, rentowność, "
                    "aktywność, modele dyskryminacyjne")
        ws['A2'].font = Font(italic=True, size=9)
        ws['A3'] = "Podmiot:"
        ws['A3'].font = self.HEADER_FONT
        ws['B3'] = self.reports[-1].dane_firmy.nazwa
        ws['A4'] = ("UWAGA: wskaźniki pełnią funkcję pomocniczą. Pełna ocena "
                    "wymaga analizy dynamicznej i kontekstu branżowego.")
        ws['A4'].font = Font(italic=True, color="666666")

        # Wzorzec wierszy = najdłuższa lista wyników (zwykle z najnowszego roku).
        wzorzec = max((w for _, w in per_report), key=len, default=[])
        if not wzorzec:
            ws['A6'] = "(nie udało się obliczyć wskaźników - brak danych)"
            return

        mapy = [{w.skrot: w for w in wyniki} for _, wyniki in per_report]
        n = len(per_report)

        header_row = 6
        naglowki = (["Wskaźnik", "Skrót"]
                    + [str(rok) for rok, _ in per_report]
                    + ["Wzór", "Optimum", "Wart. krytyczna", "Źródło"])
        for col, h in enumerate(naglowki, 1):
            c = ws.cell(row=header_row, column=col, value=h)
            c.font = self.HEADER_FONT
            c.fill = self.HEADER_FILL
            if 3 <= col <= 2 + n:
                c.alignment = Alignment(horizontal='right')

        row = header_row + 1
        biezaca_sekcja = None
        for wz in wzorzec:
            sekcja = self._indicator_section(wz.nazwa)
            if sekcja != biezaca_sekcja:
                biezaca_sekcja = sekcja
                ws.cell(row=row, column=1,
                        value=self.SECTION_LABELS.get(sekcja, sekcja)).font = self.SECTION_FONT
                row += 1

            ws.cell(row=row, column=1, value=wz.nazwa)
            ws.cell(row=row, column=2, value=wz.skrot)
            for j in range(n):
                w = mapy[j].get(wz.skrot)
                col = 3 + j
                if w is not None:
                    c = ws.cell(row=row, column=col, value=w.wartosc_str)
                    c.alignment = Alignment(horizontal='right')
                    fill = self.OCENA_FILL.get(w.ocena.value)
                    if fill:
                        c.fill = fill
            ws.cell(row=row, column=3 + n, value=wz.wzor)
            ws.cell(row=row, column=4 + n, value=wz.optimum)
            ws.cell(row=row, column=5 + n, value=wz.wartosc_krytyczna)
            ws.cell(row=row, column=6 + n, value=wz.zrodlo)
            row += 1

        # Legenda ocen.
        row += 1
        ws.cell(row=row, column=1, value="LEGENDA OCEN").font = self.SUBTITLE_FONT
        row += 1
        legenda = [
            ("optymalna", "wartość optymalna"),
            ("akceptowalna", "wartość akceptowalna"),
            ("ostrzegawcza", "wartość ostrzegawcza - wymaga uwagi"),
            ("krytyczna", "wartość krytyczna - sygnał zagrożenia niewypłacalnością"),
            ("brak_danych", "brak danych do obliczenia"),
        ]
        for ocena, opis in legenda:
            c = ws.cell(row=row, column=1, value=ocena.capitalize())
            c.fill = self.OCENA_FILL.get(ocena, PatternFill())
            ws.cell(row=row, column=2, value=opis)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=min(2 + n, 6))
            row += 1

        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 9
        for j in range(n):
            ws.column_dimensions[get_column_letter(3 + j)].width = 15
        ws.column_dimensions[get_column_letter(3 + n)].width = 50
        ws.column_dimensions[get_column_letter(4 + n)].width = 22
        ws.column_dimensions[get_column_letter(5 + n)].width = 24
        ws.column_dimensions[get_column_letter(6 + n)].width = 28
        ws.freeze_panes = ws.cell(row=header_row + 1, column=3)

    @staticmethod
    def _indicator_section(nazwa: str) -> str:
        n = nazwa.lower()
        if n.startswith("model") or "wilcox" in n:
            return "modele"
        if "rentowność" in n:
            return "rentownosc"
        if ("płynnoś" in n or "gotówk" in n or "wystarczalnoś" in n
                or n.startswith("kapitał pracując")):
            return "plynnosc"
        if "zadłużeni" in n or "udziału kapitału" in n or "pokrycia zobow" in n:
            return "zadluzenie"
        if "cykl" in n or "obrot" in n or "obrót" in n:
            return "aktywnosc"
        return "strukturalne"

    def _create_raw_sheet(self, ws):
        """Arkusz danych surowych - wszystkie pozycje z kodami, kolumny = lata."""
        getters = [
            ("Bilans-Aktywa", lambda s: s.bilans_aktywa),
            ("Bilans-Pasywa", lambda s: s.bilans_pasywa),
            ("RZiS", lambda s: s.rzis),
            ("Nota", lambda s: s.nota_podatkowa),
            ("ZmianyKapitalu", lambda s: s.zestawienie_zmian_kapital),
            ("Przeplywy", lambda s: s.rachunek_przeplywow),
        ]
        wszystkie = []  # list[tuple[label, wiersz]]
        for label, getter in getters:
            _lata, wiersze = self._merge_section(getter)
            for w in wiersze:
                wszystkie.append((label, w))

        lata = sorted({rok for _, w in wszystkie for rok in w['values']})
        naglowki = ["sekcja", "kod", "opis"] + [str(rok) for rok in lata]
        for col, h in enumerate(naglowki, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = self.HEADER_FONT
            c.fill = self.HEADER_FILL

        row = 2
        for label, w in wszystkie:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=w['kod'])
            ws.cell(row=row, column=3, value=w['opis'])
            for i, rok in enumerate(lata):
                wartosc = w['values'].get(rok)
                if wartosc is not None:
                    c = ws.cell(row=row, column=4 + i, value=float(wartosc))
                    c.number_format = self.MONEY_FORMAT
            row += 1

        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 60
        for i in range(len(lata)):
            ws.column_dimensions[get_column_letter(4 + i)].width = 16
        if row > 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(3 + len(lata))}{row - 1}"
        ws.freeze_panes = "D2"

    def _create_analytical_sheet(self, ws):
        """Arkusz danych analitycznych - format długi, wszystkie lata."""
        naglowki = ["firma", "nip", "krs", "typ_jednostki", "wersja",
                    "rok", "typ_okresu", "sekcja", "kod", "kod_pelny", "opis", "kwota"]
        for col, h in enumerate(naglowki, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = self.HEADER_FONT
            c.fill = self.HEADER_FILL

        row = 2
        for spr in self.reports:
            meta = spr.metadane
            firma = spr.dane_firmy
            rok = _rok(spr)
            for poz in spr.wszystkie_pozycje():
                kod_pelny = poz.kod_pelny(meta.typ_jednostki, meta.wersja_schematu)
                wpisy = []
                if poz.kwota_biezaca is not None:
                    wpisy.append((rok, "biezacy", poz.kwota_biezaca))
                if poz.kwota_poprzednia is not None:
                    wpisy.append((rok - 1, "poprzedni", poz.kwota_poprzednia))
                if poz.kwota_przeksztalcona is not None:
                    wpisy.append((rok - 1, "przeksztalcony", poz.kwota_przeksztalcona))
                for rok_w, typ_okresu, kwota in wpisy:
                    ws.cell(row=row, column=1, value=firma.nazwa)
                    ws.cell(row=row, column=2, value=firma.nip)
                    ws.cell(row=row, column=3, value=firma.krs or "")
                    ws.cell(row=row, column=4, value=meta.typ_jednostki)
                    ws.cell(row=row, column=5, value=meta.wersja_schematu)
                    ws.cell(row=row, column=6, value=rok_w)
                    ws.cell(row=row, column=7, value=typ_okresu)
                    ws.cell(row=row, column=8, value=poz.sekcja)
                    ws.cell(row=row, column=9, value=poz.kod)
                    ws.cell(row=row, column=10, value=kod_pelny)
                    ws.cell(row=row, column=11, value=poz.opis)
                    c = ws.cell(row=row, column=12, value=float(kwota))
                    c.number_format = self.MONEY_FORMAT
                    row += 1

        szerokosci = [40, 12, 12, 12, 8, 8, 14, 12, 25, 35, 50, 15]
        for i, w in enumerate(szerokosci, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        if row > 2:
            ws.auto_filter.ref = f"A1:L{row - 1}"
        ws.freeze_panes = "A2"

    # ----------------------------------------------------------- załączniki

    def _save_attachments(self, output_dir: Path) -> list:
        """Zapisuje załączniki binarne ze wszystkich sprawozdań.

        Każdy rok trafia do osobnego podkatalogu, aby uniknąć kolizji nazw.
        """
        saved = []
        if not any(s.zalaczniki for s in self.reports):
            return saved

        nazwa_clean = _oczysc_nazwe(self.reports[-1].dane_firmy.nazwa, limit=50)
        base_dir = Path(output_dir) / f"zalaczniki_{nazwa_clean}"

        for spr in self.reports:
            if not spr.zalaczniki:
                continue
            rok_dir = base_dir / str(_rok(spr))
            rok_dir.mkdir(parents=True, exist_ok=True)
            for i, zal in enumerate(spr.zalaczniki, 1):
                nazwa_pliku = zal.nazwa_pliku or f"zalacznik_{i}"
                output_path = rok_dir / nazwa_pliku
                if output_path.exists():
                    if "." in nazwa_pliku:
                        baza, ext = nazwa_pliku.rsplit(".", 1)
                        output_path = rok_dir / f"{baza}_{i}.{ext}"
                    else:
                        output_path = rok_dir / f"{nazwa_pliku}_{i}"
                try:
                    with open(output_path, "wb") as f:
                        f.write(zal.zawartosc)
                    saved.append(output_path)
                except IOError as e:
                    print(f"Błąd zapisu załącznika {nazwa_pliku}: {e}")
        return saved
