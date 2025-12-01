"""
Pełny konwerter dla wersji webowej.
Generuje arkusze: Bilans, RZiS, Nota podatkowa, Zestawienie zmian w kapitale,
Rachunek przepływów pieniężnych, Dane surowe, Dane analityczne.
Obsługuje załączniki i jednostkę walutową (PLN / tys. PLN).

UWAGA: Arkusz "Analiza wskaźnikowa" jest dostępny tylko w wersji lokalnej (CLI).
"""

import sys
import uuid
from pathlib import Path
from datetime import date
from decimal import Decimal
from typing import Optional, List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Dodaj ścieżkę do głównego modułu src
# W Docker: /app/app/converter_simple.py -> /app/src
# Lokalnie: web/app/converter_simple.py -> src
SRC_PATH = Path(__file__).parent.parent / "src"  # dla Docker (/app/src)
SRC_PATH_LOCAL = Path(__file__).parent.parent.parent / "src"  # dla lokalnego dev

for path in [SRC_PATH, SRC_PATH_LOCAL]:
    if path.exists() and str(path) not in sys.path:
        sys.path.insert(0, str(path))
        break

from parser import SFParser
from models import Sprawozdanie, PozycjaFinansowa, Zalacznik


class SimpleXLSXConverter:
    """Pełny konwerter webowy - wszystkie sekcje sprawozdania finansowego."""

    HEADER_FONT = Font(bold=True)
    HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    TITLE_FONT = Font(bold=True, size=12)
    MONEY_FORMAT = '#,##0.00'
    MONEY_FORMAT_TYS = '#,##0.00" tys."'
    DATE_FORMAT = 'YYYY-MM-DD'

    def convert(self, xml_path: Path, output_path: Path, attachments_dir: Optional[Path] = None) -> dict:
        """
        Konwertuje plik XML do pełnego XLSX z wszystkimi sekcjami sprawozdania.

        Args:
            xml_path: Ścieżka do pliku XML
            output_path: Ścieżka do pliku wyjściowego XLSX
            attachments_dir: Katalog na zapisanie załączników (opcjonalnie)

        Returns:
            dict z metadanymi sprawozdania
        """
        # Parsuj XML
        parser = SFParser()
        spr = parser.parse(xml_path)

        # Określ jednostkę walutową
        self.jednostka_walutowa = spr.metadane.jednostka_walutowa
        self.money_format = self.MONEY_FORMAT_TYS if self.jednostka_walutowa == "tys. PLN" else self.MONEY_FORMAT

        # Utwórz workbook
        wb = Workbook()

        # Arkusz 1: Bilans
        ws_bilans = wb.active
        ws_bilans.title = "Bilans"
        self._create_bilans_sheet(ws_bilans, spr)

        # Arkusz 2: RZiS
        wariant = "w.por." if spr.metadane.wariant_rzis == "porownawczy" else "w.kalk."
        ws_rzis = wb.create_sheet(f"RZiS ({wariant})")
        self._create_rzis_sheet(ws_rzis, spr)

        # Arkusz 3: Nota podatkowa (jeśli dostępna)
        if spr.nota_podatkowa:
            ws_nota = wb.create_sheet("Nota podatkowa")
            self._create_nota_sheet(ws_nota, spr)

        # Arkusz 4: Zestawienie zmian w kapitale własnym (jeśli dostępne)
        if spr.zestawienie_zmian_kapital:
            ws_kapital = wb.create_sheet("Zest. zmian w kapitale")
            self._create_kapital_sheet(ws_kapital, spr)

        # Arkusz 5: Rachunek przepływów pieniężnych (jeśli dostępny)
        if spr.rachunek_przeplywow:
            wariant_przep = "bezp." if spr.wariant_przeplywow == "bezposredni" else "pośr."
            ws_przeplywy = wb.create_sheet(f"Rach. przepływów ({wariant_przep})")
            self._create_przeplywy_sheet(ws_przeplywy, spr)

        # Zapisz
        wb.save(output_path)

        # Zapisz załączniki jeśli są i podano katalog
        saved_attachments = []
        if attachments_dir and spr.zalaczniki:
            saved_attachments = self._save_attachments(spr.zalaczniki, attachments_dir)

        # Generuj czytelną nazwę pliku
        readable_filename = self._generate_filename(spr)

        # Generuj dane do podglądu
        preview_data = self._generate_preview(spr)

        # Zwróć metadane
        return {
            "company_name": spr.dane_firmy.nazwa,
            "company_nip": spr.dane_firmy.nip,
            "company_krs": spr.dane_firmy.krs,
            "company_regon": spr.dane_firmy.regon,
            "company_address": spr.dane_firmy.adres_pelny(),
            "entity_type": spr.metadane.typ_jednostki,
            "period_from": spr.metadane.okres_od,
            "period_to": spr.metadane.okres_do,
            "data_sporzadzenia": spr.metadane.data_sporzadzenia,
            "jednostka_walutowa": spr.metadane.jednostka_walutowa,
            "wariant_rzis": spr.metadane.wariant_rzis,
            "output_filename": readable_filename,
            "attachments": saved_attachments,
            "preview": preview_data,
        }

    def _save_attachments(self, zalaczniki: List[Zalacznik], attachments_dir: Path) -> List[Dict[str, Any]]:
        """Zapisuje załączniki do plików i zwraca ich metadane."""
        saved = []
        for zal in zalaczniki:
            # Generuj unikalny identyfikator dla pliku
            unique_id = str(uuid.uuid4())
            # Zachowaj oryginalne rozszerzenie
            ext = zal.rozszerzenie() or "bin"
            safe_filename = f"{unique_id}.{ext}"
            file_path = attachments_dir / safe_filename

            # Zapisz plik
            file_path.write_bytes(zal.zawartosc)

            saved.append({
                "id": unique_id,
                "original_name": zal.nazwa_pliku,
                "safe_filename": safe_filename,
                "path": str(file_path),
                "size_kb": round(zal.rozmiar_kb(), 2),
                "extension": ext,
                "sekcja": zal.sekcja,
                "opis": zal.opis,
            })

        return saved

    def _generate_preview(self, spr: Sprawozdanie) -> Dict[str, Any]:
        """Generuje dane do podglądu konwersji."""
        # Dane ogólne
        general_info = {
            "firma": spr.dane_firmy.nazwa,
            "nip": spr.dane_firmy.nip,
            "krs": spr.dane_firmy.krs,
            "regon": spr.dane_firmy.regon,
            "adres": spr.dane_firmy.adres_pelny(),
            "okres": f"{spr.metadane.okres_od} - {spr.metadane.okres_do}",
            "typ_jednostki": spr.metadane.typ_jednostki,
            "jednostka_walutowa": spr.metadane.jednostka_walutowa,
            "wariant_rzis": "porównawczy" if spr.metadane.wariant_rzis == "porownawczy" else "kalkulacyjny",
        }

        # Weryfikacja sum
        weryfikacja = None
        if spr.weryfikacja:
            weryfikacja = {
                "aktywa_biezacy": self._format_kwota(spr.weryfikacja.aktywa_razem_biezacy),
                "pasywa_biezacy": self._format_kwota(spr.weryfikacja.pasywa_razem_biezacy),
                "zgodnosc": spr.weryfikacja.aktywa_rowne_pasywom_biezacy,
            }

        # Pierwsze 20 wierszy bilansu (aktywa + pasywa razem)
        bilans_preview = []
        all_bilans = spr.bilans_aktywa + spr.bilans_pasywa
        for i, poz in enumerate(all_bilans[:20]):
            bilans_preview.append({
                "lp": i + 1,
                "pozycja": poz.opis,
                "kwota_biezaca": self._format_kwota(poz.kwota_biezaca),
                "kwota_poprzednia": self._format_kwota(poz.kwota_poprzednia),
                "poziom": poz.poziom,
            })

        # Informacje o dostępnych sekcjach
        sekcje = {
            "bilans": len(all_bilans) > 0,
            "rzis": len(spr.rzis) > 0,
            "nota_podatkowa": spr.nota_podatkowa is not None and len(spr.nota_podatkowa) > 0,
            "zestawienie_zmian_kapital": spr.zestawienie_zmian_kapital is not None and len(spr.zestawienie_zmian_kapital) > 0,
            "rachunek_przeplywow": spr.rachunek_przeplywow is not None and len(spr.rachunek_przeplywow) > 0,
        }

        return {
            "general": general_info,
            "weryfikacja": weryfikacja,
            "bilans": bilans_preview,
            "bilans_total": len(all_bilans),
            "rzis_total": len(spr.rzis),
            "nota_total": len(spr.nota_podatkowa) if spr.nota_podatkowa else 0,
            "kapital_total": len(spr.zestawienie_zmian_kapital) if spr.zestawienie_zmian_kapital else 0,
            "przeplywy_total": len(spr.rachunek_przeplywow) if spr.rachunek_przeplywow else 0,
            "wariant_przeplywow": spr.wariant_przeplywow if spr.rachunek_przeplywow else None,
            "sekcje": sekcje,
            "attachments_count": len(spr.zalaczniki),
        }

    def _format_kwota(self, kwota: Optional[Decimal]) -> str:
        """Formatuje kwotę z polskim formatem (przecinek jako separator dziesiętny)."""
        if kwota is None:
            return "-"
        # Formatuj z 2 miejscami po przecinku i spacjami jako separatorami tysięcy
        formatted = f"{kwota:,.2f}".replace(",", " ").replace(".", ",")
        return formatted

    def _generate_filename(self, spr: Sprawozdanie) -> str:
        """Generuje czytelną nazwę pliku wyjściowego."""
        meta = spr.metadane
        firma = spr.dane_firmy.nazwa

        # Usuń znaki niedozwolone w nazwach plików Windows
        niedozwolone = '<>:"/\\|?*'
        firma_clean = "".join(c for c in firma if c not in niedozwolone)

        # Ogranicz długość nazwy firmy
        if len(firma_clean) > 50:
            firma_clean = firma_clean[:50]

        # Format: SF_2023_NazwaFirmy.xlsx
        return f"SF_{meta.okres_do.year}_{firma_clean}.xlsx"

    def _create_bilans_sheet(self, ws, spr: Sprawozdanie):
        """Tworzy arkusz Bilans."""
        meta = spr.metadane
        firma = spr.dane_firmy
        weryfikacja = spr.weryfikacja
        jednostka = meta.jednostka_walutowa

        # Nagłówek
        ws['A1'] = "BILANS"
        ws['A1'].font = self.TITLE_FONT

        # Dane firmy
        ws['A2'] = "Firma:"
        ws['B2'] = firma.nazwa

        ws['A3'] = "NIP:"
        ws['B3'] = firma.nip

        if firma.krs:
            ws['A4'] = "KRS:"
            ws['B4'] = firma.krs

        # Okresy
        row = 6
        ws[f'A{row}'] = "Okres sprawozdawczy:"
        ws[f'B{row}'] = f"{meta.okres_od} - {meta.okres_do}"

        row += 1
        ws[f'A{row}'] = "Typ jednostki:"
        ws[f'B{row}'] = meta.typ_jednostki

        row += 1
        ws[f'A{row}'] = "Jednostka walutowa:"
        ws[f'B{row}'] = jednostka

        # Weryfikacja sum
        row += 2
        ws[f'A{row}'] = "Weryfikacja sum:"
        if weryfikacja:
            if weryfikacja.aktywa_rowne_pasywom_biezacy:
                ws[f'B{row}'] = "OK (Aktywa = Pasywa)"
            else:
                ws[f'B{row}'] = "BŁĄD: Aktywa ≠ Pasywa"
                ws[f'B{row}'].font = Font(color="FF0000", bold=True)

        # Nagłówki kolumn
        row += 2
        ws[f'A{row}'] = "Pozycja"
        ws[f'B{row}'] = f"Rok {meta.okres_do.year} ({jednostka})"
        ws[f'C{row}'] = f"Rok {meta.okres_do.year - 1} ({jednostka})"
        ws[f'D{row}'] = f"Rok {meta.okres_do.year - 1} przekształcone ({jednostka})"

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = self.HEADER_FONT
            ws[f'{col}{row}'].fill = self.HEADER_FILL

        # AKTYWA
        row += 2
        ws[f'A{row}'] = "AKTYWA"
        ws[f'A{row}'].font = self.TITLE_FONT
        row += 1

        for poz in spr.bilans_aktywa:
            row = self._write_position_row(ws, row, poz, has_przeksztalcona=True)

        # PASYWA
        row += 2
        ws[f'A{row}'] = "PASYWA"
        ws[f'A{row}'].font = self.TITLE_FONT
        row += 1

        for poz in spr.bilans_pasywa:
            row = self._write_position_row(ws, row, poz, has_przeksztalcona=True)

        # Szerokości kolumn
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22

    def _create_rzis_sheet(self, ws, spr: Sprawozdanie):
        """Tworzy arkusz RZiS."""
        meta = spr.metadane
        firma = spr.dane_firmy
        jednostka = meta.jednostka_walutowa

        # Nagłówek
        wariant_nazwa = "WARIANT PORÓWNAWCZY" if meta.wariant_rzis == "porownawczy" else "WARIANT KALKULACYJNY"
        ws['A1'] = f"RACHUNEK ZYSKÓW I STRAT ({wariant_nazwa})"
        ws['A1'].font = self.TITLE_FONT

        # Dane firmy
        ws['A2'] = "Firma:"
        ws['B2'] = firma.nazwa

        ws['A3'] = "Okres:"
        ws['B3'] = f"{meta.okres_od} - {meta.okres_do}"

        ws['A4'] = "Jednostka walutowa:"
        ws['B4'] = jednostka

        # Nagłówki kolumn
        row = 6
        ws[f'A{row}'] = "Pozycja"
        ws[f'B{row}'] = f"Rok {meta.okres_do.year} ({jednostka})"
        ws[f'C{row}'] = f"Rok {meta.okres_do.year - 1} ({jednostka})"
        ws[f'D{row}'] = f"Rok {meta.okres_do.year - 1} przekształcone ({jednostka})"

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = self.HEADER_FONT
            ws[f'{col}{row}'].fill = self.HEADER_FILL

        # Pozycje RZiS
        row += 1
        for poz in spr.rzis:
            row = self._write_position_row(ws, row, poz, has_przeksztalcona=True)

        # Szerokości kolumn
        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22

    def _create_nota_sheet(self, ws, spr: Sprawozdanie):
        """Tworzy arkusz Nota podatkowa.

        UWAGA: Schemat noty podatkowej (XML) nie zawiera trzeciej kolumny
        (przekształcone dane porównawcze). Kolumna D jest dodawana dla
        spójności formatu, ale pozostaje pusta.
        """
        meta = spr.metadane
        firma = spr.dane_firmy
        jednostka = meta.jednostka_walutowa

        # Nagłówek
        ws['A1'] = "NOTA PODATKOWA (Dodatkowe Informacje i Objaśnienia)"
        ws['A1'].font = self.TITLE_FONT

        ws['A2'] = "Firma:"
        ws['B2'] = firma.nazwa

        ws['A3'] = "Okres:"
        ws['B3'] = f"{meta.okres_od} - {meta.okres_do}"

        # Nagłówki kolumn
        row = 5
        ws[f'A{row}'] = "Pozycja"
        ws[f'B{row}'] = f"Rok {meta.okres_do.year} ({jednostka})"
        ws[f'C{row}'] = f"Rok {meta.okres_do.year - 1} ({jednostka})"
        ws[f'D{row}'] = f"Rok {meta.okres_do.year - 1} przekształcone ({jednostka})"

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = self.HEADER_FONT
            ws[f'{col}{row}'].fill = self.HEADER_FILL

        # Pozycje (nota podatkowa nie ma danych przekształconych w schemacie XML)
        row += 1
        for poz in spr.nota_podatkowa:
            row = self._write_position_row(ws, row, poz, has_przeksztalcona=False)

        # Formatowanie kolumn
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22

    def _create_kapital_sheet(self, ws, spr: Sprawozdanie):
        """Tworzy arkusz Zestawienie zmian w kapitale własnym."""
        meta = spr.metadane
        firma = spr.dane_firmy
        jednostka = meta.jednostka_walutowa

        # Nagłówek
        ws['A1'] = "ZESTAWIENIE ZMIAN W KAPITALE (FUNDUSZU) WŁASNYM"
        ws['A1'].font = self.TITLE_FONT

        ws['A2'] = "Firma:"
        ws['B2'] = firma.nazwa

        ws['A3'] = "Okres:"
        ws['B3'] = f"{meta.okres_od} - {meta.okres_do}"

        # Nagłówki kolumn
        row = 5
        ws[f'A{row}'] = "Pozycja"
        ws[f'B{row}'] = f"Rok {meta.okres_do.year} ({jednostka})"
        ws[f'C{row}'] = f"Rok {meta.okres_do.year - 1} ({jednostka})"
        ws[f'D{row}'] = f"Rok {meta.okres_do.year - 1} przekształcone ({jednostka})"

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = self.HEADER_FONT
            ws[f'{col}{row}'].fill = self.HEADER_FILL

        # Pozycje
        row += 1
        for poz in spr.zestawienie_zmian_kapital:
            row = self._write_position_row(ws, row, poz, has_przeksztalcona=True)

        # Formatowanie kolumn
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22

    def _create_przeplywy_sheet(self, ws, spr: Sprawozdanie):
        """Tworzy arkusz Rachunek przepływów pieniężnych."""
        meta = spr.metadane
        firma = spr.dane_firmy
        jednostka = meta.jednostka_walutowa

        # Nagłówek
        wariant_nazwa = "METODA BEZPOŚREDNIA" if spr.wariant_przeplywow == "bezposredni" else "METODA POŚREDNIA"
        ws['A1'] = f"RACHUNEK PRZEPŁYWÓW PIENIĘŻNYCH ({wariant_nazwa})"
        ws['A1'].font = self.TITLE_FONT

        ws['A2'] = "Firma:"
        ws['B2'] = firma.nazwa

        ws['A3'] = "Okres:"
        ws['B3'] = f"{meta.okres_od} - {meta.okres_do}"

        # Nagłówki kolumn
        row = 5
        ws[f'A{row}'] = "Pozycja"
        ws[f'B{row}'] = f"Rok {meta.okres_do.year} ({jednostka})"
        ws[f'C{row}'] = f"Rok {meta.okres_do.year - 1} ({jednostka})"
        ws[f'D{row}'] = f"Rok {meta.okres_do.year - 1} przekształcone ({jednostka})"

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = self.HEADER_FONT
            ws[f'{col}{row}'].fill = self.HEADER_FILL

        # Pozycje
        row += 1
        for poz in spr.rachunek_przeplywow:
            row = self._write_position_row(ws, row, poz, has_przeksztalcona=True)

        # Formatowanie kolumn
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22

    def _write_position_row(self, ws, row: int, poz: PozycjaFinansowa, has_przeksztalcona: bool = True) -> int:
        """Zapisuje wiersz pozycji finansowej.

        Args:
            ws: Arkusz
            row: Numer wiersza
            poz: Pozycja finansowa
            has_przeksztalcona: Czy arkusz ma kolumnę z danymi przekształconymi

        Returns:
            Numer następnego wiersza
        """
        # Wcięcie
        indent = "  " * poz.poziom
        ws[f'A{row}'] = f"{indent}{poz.opis}"

        # Pogrubienie dla pozycji głównych
        if poz.poziom == 0:
            ws[f'A{row}'].font = Font(bold=True)

        # Kwoty
        if poz.kwota_biezaca is not None:
            cell = ws[f'B{row}']
            cell.value = float(poz.kwota_biezaca)
            cell.number_format = self.money_format
            cell.alignment = Alignment(horizontal='right')

        if poz.kwota_poprzednia is not None:
            cell = ws[f'C{row}']
            cell.value = float(poz.kwota_poprzednia)
            cell.number_format = self.money_format
            cell.alignment = Alignment(horizontal='right')

        # Kwota przekształcona (trzecia kolumna)
        if has_przeksztalcona and poz.kwota_przeksztalcona is not None:
            cell = ws[f'D{row}']
            cell.value = float(poz.kwota_przeksztalcona)
            cell.number_format = self.money_format
            cell.alignment = Alignment(horizontal='right')

        return row + 1


def convert_file(xml_path: str, output_path: str, attachments_dir: Optional[str] = None) -> dict:
    """
    Funkcja pomocnicza do konwersji.

    Args:
        xml_path: Ścieżka do pliku XML
        output_path: Ścieżka do pliku XLSX
        attachments_dir: Katalog na załączniki (opcjonalnie)

    Returns:
        dict z metadanymi sprawozdania
    """
    converter = SimpleXLSXConverter()
    attach_path = Path(attachments_dir) if attachments_dir else None
    return converter.convert(Path(xml_path), Path(output_path), attach_path)
